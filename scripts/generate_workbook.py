import openpyxl
from openpyxl.styles import PatternFill, Font
import os # <--- 追加

def create_trading_hub_workbook(filename="trading_hub.xlsm"):
    """
    取引ハブとなるExcelワークブック(trading_hub.xlsm)を生成します。
    このファイルには、データ取得用と注文執行用のシートが含まれます。
    """
    # ワークブックを作成
    wb = openpyxl.Workbook()

    # スタイルを定義
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    trigger_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    # ==========================================================================
    # 1. 「リアルタイムデータ」シートの作成
    # ==========================================================================
    data_sheet = wb.active
    data_sheet.title = "リアルタイムデータ"

    # ヘッダーを設定
    data_headers = ["銘柄コード", "現在値", "始値", "高値", "安値", "出来高"]
    data_sheet.append(data_headers)
    for cell in data_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # サンプルデータとMS2のRSS関数を設定
    data_sheet["A2"] = 9984
    data_sheet["B2"] = '=MS2.RSS(A2,"現在値")'
    data_sheet["C2"] = '=MS2.RSS(A2,"始値")'
    data_sheet["D2"] = '=MS2.RSS(A2,"高値")'
    data_sheet["E2"] = '=MS2.RSS(A2,"安値")'
    data_sheet["F2"] = '=MS2.RSS(A2,"出来高")'
    
    data_sheet["A3"] = 7203
    data_sheet["B3"] = '=MS2.RSS(A3,"現在値")'
    data_sheet["C3"] = '=MS2.RSS(A3,"始値")'
    data_sheet["D3"] = '=MS2.RSS(A3,"高値")'
    data_sheet["E3"] = '=MS2.RSS(A3,"安値")'
    data_sheet["F3"] = '=MS2.RSS(A3,"出来高")'

    # 口座情報を設定
    data_sheet["A11"] = "買付余力"
    data_sheet["B11"] = '=MS2.ACCOUNT("買付余力")'
    data_sheet["A11"].font = header_font
    
    print(f"シート '{data_sheet.title}' を作成しました。")

    # ==========================================================================
    # 2. 「注文」シートの作成
    # ==========================================================================
    order_sheet = wb.create_sheet(title="注文")

    # ヘッダーを設定
    order_headers = ["項目", "入力 (Python →)", "結果 (→ Python)", "同期トリガー"]
    order_sheet.append(order_headers)
    for cell in order_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 各項目ラベルと初期値を設定
    order_layout = {
        "A2": "銘柄コード", "A3": "売買区分", "A4": "数量",
        "A5": "注文種別", "A6": "指値価格",
        "C2": "処理ステータス", "C3": "注文ID", "C4": "メッセージ",
        "D2": "実行ID", "D3": "完了ID"
    }
    for cell_ref, value in order_layout.items():
        order_sheet[cell_ref] = value
        order_sheet[cell_ref].font = header_font
        
    # トリガーの初期値を設定
    order_sheet["D2"].offset(row=0, column=1).value = 0 # E2セル
    order_sheet["D3"].offset(row=0, column=1).value = 0 # E3セル

    # セルの色分け
    for i in range(2, 7):
        order_sheet[f"B{i}"].fill = input_fill
    for i in range(2, 5):
        order_sheet[f"C{i}"].fill = output_fill
    order_sheet["D2"].fill = trigger_fill
    order_sheet["D3"].fill = trigger_fill
    order_sheet["E2"].fill = trigger_fill
    order_sheet["E3"].fill = trigger_fill

    print(f"シート '{order_sheet.title}' を作成しました。")
    
    # ==========================================================================
    # 3. ファイルを保存
    # ==========================================================================
    
    # --- ▼▼▼ 変更箇所 ▼▼▼ ---
    directory = "external"
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"ディレクトリ '{directory}' を作成しました。")
        
    save_path = os.path.join(directory, filename)
    # --- ▲▲▲ 変更箇所 ▲▲▲ ---

    try:
        # --- ▼▼▼ 変更箇所 ▼▼▼ ---
        wb.save(save_path)
        print(f"\nワークブック '{save_path}' を正常に生成しました。")
        # --- ▲▲▲ 変更箇所 ▲▲▲ ---
        print("注意: このスクリプトはVBAマクロを含みません。Excelを開き、手動でVBAエディタからマクロを追加してください。")
    except Exception as e:
        print(f"エラー: ファイルの保存に失敗しました。 - {e}")


if __name__ == '__main__':
    create_trading_hub_workbook()